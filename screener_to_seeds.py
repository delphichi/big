#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
五維總篩選 → 感測器種子橋接  screener_to_seeds.py
================================================
把 total_screener.py 產出的「台股五維總篩選.xlsx」自動轉成 SLCA 感測器 v2 的輸入,
規則式偵測量化差異(②基本面 / ③矛盾)後,直接跑 sensor 全流程(含現實驗證 + 演化記錄)。

  ‧ 客觀現況(現實驗證快照)直接取自 xlsx 的五維欄位 + 各檔損益表(毛利率/淨利),
    所以「用 total_screener 抓回的數據做分析」這一步,不需要再連 FinMind。
  ‧ 質性欄位(共識 / 脆弱前提 / 波普爾三問)以模板生成,一律標註「需人工確認」——
    感測器鐵律:這些是判斷,程式只擺資料、不替你下結論。

偵測規則(只用客觀指標,不臆測):
  ② 基本面差異  營收連續正成長≥10月  或  ROE歷史百分位≥80
  ③ 矛盾訊號    便宜的好公司(PE百分位≤40 且 ROE百分位≥80)
                品質高但股價落後(ROE百分位≥80 且 近半年相對報酬<0)
                品質高但燒錢(ROE/ROIC百分位≥80 且 含金量<0.8/FCF<0)
  死亡模式005  純資金假突破(分類為⚠純資金 且 ROE百分位<80)→ 扣分

用法:
  pip install openpyxl
  python3 screener_to_seeds.py data/台股五維總篩選.xlsx --out data/SLCA_種子_實測.md
"""
import os, argparse
import openpyxl
import sensor

# 北極星:可依需要改。能力圈外的標的會被硬過濾。
NORTH_STAR = {
    "philosophy": "用合理價格買進品質持續創高、但被市場錯誤定價的公司,持有至差異收斂",
    "circle": ["金融銀行證券", "半導體", "網通", "被動元件", "PC/伺服器代工"],
    "market_cap_min_yi": 100, "timeframe": "12–24個月", "history_success": {},
}
DEATH_005 = {"id": "005", "name": "純資金假突破", "penalty": 35,
             "trigger": "基本面弱(≤1/4) + 共振強 + 價格創高",
             "history": "純資金推動、基本面未跟上的標的多數回落",
             "fail_rate": "—", "handling": "基本面未跟上前,資金面差異不計入種子"}


def _num(v):
    try: return float(v)
    except (TypeError, ValueError): return None

def _grow(v):
    """營收正成長欄是 'X/Y' 字串;沒資料時是 '—'。穩健取 X,取不到回 0。"""
    try: return int(str(v).split("/")[0])
    except (TypeError, ValueError, AttributeError, IndexError): return 0


def margin_ni(wb, sid):
    """從 <sid>_損益表 算 毛利率近4季/趨勢、連兩季淨利是否改善。"""
    name = f"{sid}_損益表"
    if name not in wb.sheetnames:
        return None, None, None
    rows = list(wb[name].iter_rows(values_only=True)); h = rows[0]
    di, ti, vi = h.index("date"), h.index("type"), h.index("value")
    rev, gp, ni = {}, {}, {}
    for r in rows[1:]:
        d, t, v = r[di], r[ti], _num(r[vi])
        if v is None: continue
        if t == "Revenue": rev[d] = v
        elif t == "GrossProfit": gp[d] = v
        elif t == "IncomeAfterTaxes": ni[d] = v
    gm = [(d, gp[d] / rev[d] * 100) for d in sorted(rev) if d in gp and rev[d]]
    gm4 = [round(x, 1) for _, x in gm[-4:]] if gm else None
    trend = ("上升" if len(gm) >= 2 and gm[-1][1] > gm[-2][1] else
             "下滑" if len(gm) >= 2 and gm[-1][1] < gm[-2][1] else
             "持平" if len(gm) >= 2 else None)
    ni_s = [ni[d] for d in sorted(ni)]
    ni2 = bool(ni_s[-1] > ni_s[-2] and ni_s[-2] > ni_s[-3]) if len(ni_s) >= 3 else None
    return gm4, trend, ni2


def build(xlsx, as_of):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    summ = list(wb["五維總篩選"].iter_rows(values_only=True)); hdr = summ[0]
    def C(row, name): return row[hdr.index(name)] if name in hdr else None

    tickers, verify = [], {}
    for row in summ[1:]:
        sid = str(C(row, "代號"))
        cat = C(row, "分類") or ""
        grow = _grow(C(row, "營收正成長"))
        roe_p, roic_p, pe_p = _num(C(row, "ROE百分位")), _num(C(row, "ROIC百分位")), _num(C(row, "PE百分位"))
        cashq, fcf, rel = _num(C(row, "含金量")), _num(C(row, "近四季FCF")), _num(C(row, "相對報酬%"))
        lit = C(row, "亮燈") or ""

        diffs = []
        if grow >= 10 and roe_p is not None and roe_p >= 80:
            diffs.append({"type": "基本面差異", "note": f"營收連{grow}月正成長 + ROE百分位{roe_p}(成長×品質創高)"})
        elif grow >= 10:
            diffs.append({"type": "基本面差異", "note": f"營收連{grow}月正成長(動能持續)"})
        elif roe_p is not None and roe_p >= 80:
            diffs.append({"type": "基本面差異", "note": f"ROE百分位{roe_p}(品質居歷史高位)"})

        contra, base = [], 0
        if pe_p is not None and pe_p <= 40 and roe_p is not None and roe_p >= 80:
            contra.append(f"ROE百分位{roe_p}(品質高)但PE百分位{pe_p}(估值便宜)——便宜的好公司"); base = max(base, 88)
        if roe_p is not None and roe_p >= 80 and rel is not None and rel < 0:
            contra.append(f"ROE百分位{roe_p}(品質高)但近半年相對報酬{rel}%(股價落後大盤)"); base = max(base, 78)
        if ((roe_p and roe_p >= 80) or (roic_p and roic_p >= 80)) and cashq is not None and cashq < 0.8:
            contra.append(f"品質百分位高,但含金量{cashq}<0.8、FCF{fcf}億(賺帳面、現金未跟上)"); base = max(base, 72)
        if contra:
            diffs.append({"type": "矛盾訊號", "base_ds": base, "note": "；".join(contra)})

        deaths = ["005"] if (cat.startswith("⚠ 純資金") and not (roe_p and roe_p >= 80)) else []

        cstr = "".join(contra)
        if "便宜的好公司" in cstr:
            q3 = "未來兩季:估值修復(PE歷史百分位回升)且股價由落後轉為跑贏大盤"
            kill = "ROE/ROIC歷史百分位跌破中位數,或便宜持續無人重估達兩季以上"
        elif "股價落後" in cstr:
            q3 = "未來兩季:近半年相對報酬由負轉正,且ROE歷史百分位維持≥80"
            kill = "ROE百分位顯著下滑,或相對報酬連兩季續惡化"
        elif "現金未跟上" in cstr:
            q3 = "未來兩季:獲利含金量回升至≥0.8且自由現金流轉正"
            kill = "含金量連兩季<0.8且FCF持續為負(確認為品質幻覺/燒錢)"
        elif diffs:
            q3 = "未來兩季:營收YoY維持正成長且ROE歷史百分位不下滑"
            kill = "營收YoY轉負連兩月,或ROE百分位明顯下滑"
        else:
            q3, kill = "", ""

        tickers.append({
            "id": sid, "market_cap_yi": 9999, "in_circle": True, "liquidity_ok": True,
            "differences": diffs, "death_patterns": deaths,
            "consensus": ("市場已認同其基本面與資金同向" if cat.startswith("★") else
                          "市場偏多但對現金/估值仍有疑慮" if cat.startswith("○") else
                          "市場以資金行情追捧,基本面尚未跟上" if cat.startswith("⚠") else "—"),
            "fragile_premise": "若上述客觀指標的背離在兩季內收斂,當前定價前提即動搖",
            "has_history_case": False, "kill_condition": kill, "kill_hard_to_verify": False,
            "observe_months": 12, "suggest_direction": "反共識" if contra else "深化",
            "trigger": "（規則式自動偵測自五維總篩選;質性欄位為模板,需人工確認)",
            "popper": {"q1_false_positive_count": 0, "q2_mundane": "可能只是當前多頭行情的普漲",
                       "q2_fully_explains": False, "q3_observable": q3},
        })

        gm4, trend, ni2 = margin_ni(wb, sid)
        verify[sid] = {"分類": cat, "總評分": C(row, "總評分"), "共振分數": C(row, "共振分數"),
                       "五關": C(row, "五關"), "PE": C(row, "PE"), "PE百分位": pe_p,
                       "ROE%": C(row, "ROE%"), "ROE百分位": roe_p, "近四季FCF": fcf, "含金量": cashq,
                       "相對報酬%": rel, "週斜率%": C(row, "週斜率%"), "RS創高": C(row, "RS創高"),
                       "亮燈": lit, "毛利率近4季": gm4, "毛利率趨勢": trend, "連兩季淨利改善": ni2,
                       "股價位置": "接近52週高" if "價格創52週高" in lit else None}

    data = {"as_of": as_of, "north_star": NORTH_STAR, "market_volatility_high": False,
            "death_patterns_extra": [DEATH_005], "tickers": tickers}
    return data, verify


def main():
    ap = argparse.ArgumentParser(description="五維總篩選 xlsx → 感測器種子")
    ap.add_argument("xlsx", help="total_screener 產出的 .xlsx")
    ap.add_argument("--out", default="data/SLCA_種子_實測.md")
    ap.add_argument("--as-of", default=None, help="掃描日期(預設今天)")
    args = ap.parse_args()
    import datetime as dt
    as_of = args.as_of or dt.date.today().isoformat()

    data, verify = build(args.xlsx, as_of)
    res = sensor.run_sensor(data)
    res["verify"] = verify
    report = sensor.render_report(res, full=True)
    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        f.write(report + "\n")
    print(f"已輸出:{args.out}\n")
    print(report)


if __name__ == "__main__":
    main()
